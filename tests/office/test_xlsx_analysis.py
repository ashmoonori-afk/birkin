from __future__ import annotations

import hashlib
from datetime import date
from pathlib import Path
from typing import cast

from docx import Document
from openpyxl import Workbook

from birkin.office.service import DocumentService


def _artifact(path: Path) -> dict[str, str]:
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return {"uri": str(path), "content_hash": digest}


def _workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Period", "Category", "Amount", "Rate", "TextNumber", "Formula"])
    rows = [
        [date(2026, 1, 1), "A", 100, 0.10, "200", "=C2*2"],
        [date(2026, 1, 1), "A", 50, 0.20, "300", "=C3*2"],
        [date(2026, 2, 1), "A", 130, 0.15, "400", "=C4*2"],
        [date(2026, 2, 1), "B", 40, 0.05, "500", "=C5*2"],
        [date(2026, 2, 1), "B", 40, 0.05, "500", "=C5*2"],
        [date(2026, 2, 1), "Hidden", 999, 0.99, "999", "=C7*2"],
    ]
    for row in rows:
        sheet.append(row)
    for row in range(2, 8):
        sheet.cell(row, 3).number_format = "$#,##0.00"
        sheet.cell(row, 4).number_format = "0%"
    sheet.row_dimensions[7].hidden = True
    workbook.save(path)
    return path


def test_xlsx_review_aggregates_with_cell_evidence_and_explicit_policies(tmp_path: Path) -> None:
    service = DocumentService(tmp_path)
    source = _workbook(tmp_path / "sales.xlsx")

    result = service.analyze_workbook(
        _artifact(source),
        sheet="Sales",
        cell_range="A1:F7",
        group_by="Category",
        value_column="Amount",
        compare_by="Period",
    )

    assert result["source_sha256"] == _artifact(source)["content_hash"]
    assert result["selection"]["hidden_rows_excluded"] == [7]
    assert result["profile"]["duplicates"] == [{"row": 6, "duplicate_of": 5}]
    aggregate = result["aggregate"]
    assert aggregate["sum"] == 360
    assert aggregate["evidence"] == ["C2", "C3", "C4", "C5", "C6"]
    assert aggregate["groups"] == [
        {"key": "A", "sum": 280, "evidence": ["C2", "C3", "C4"]},
        {"key": "B", "sum": 80, "evidence": ["C5", "C6"]},
    ]
    assert aggregate["comparison"]["delta"] == 60
    assert result["calculation"] == {
        "performed": False,
        "formula_cache": {"missing": 5},
        "status": "not_recalculated",
    }
    assert result["policies"]["numeric_strings"].startswith("kept as text")

    created = service.create_document(
        format="docx",
        content=cast("dict[str, object]", result["report_content"]),
        output_name="sales-review.docx",
    )
    assert Document(created["draft_artifact"]["uri"]).paragraphs[0].text == "Sales 데이터 검토"
